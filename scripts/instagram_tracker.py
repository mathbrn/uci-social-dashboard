#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UCI Social Dashboard — Instagram Followers Tracker
===================================================
Scrape le nombre de followers Instagram des équipes et courses UCI WorldTour
via Playwright (gratuit, 100% local), et :
  1. Sauvegarde l'historique dans historique_followers.xlsx (fichier local)
  2. Pousse les données dans le Gist GitHub public du dashboard (optionnel)

Usage de base :
    python instagram_tracker.py

Pour activer la synchronisation avec le dashboard, créez deux variables
d'environnement avant de lancer le script :
    set UCI_GIST_ID=<id-du-gist>
    set UCI_GIST_TOKEN=<votre-pat-github-scope-gist>
    python instagram_tracker.py

Prérequis :
    pip install playwright openpyxl
    python -m playwright install chromium
"""

import json
import os
import re
import sys
import time
import urllib.request
import urllib.error
from datetime import datetime, timezone
from pathlib import Path

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("❌ Playwright non installé. Lancez : pip install playwright && python -m playwright install chromium")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("❌ openpyxl non installé. Lancez : pip install openpyxl")
    sys.exit(1)

# ═══════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════

SCRIPT_DIR = Path(__file__).resolve().parent
EXCEL_FILE = SCRIPT_DIR / "historique_followers.xlsx"
GIST_FILENAME = "uci-dashboard-data.json"

# Sync Gist (optionnel) — définir via variables d'environnement
GIST_ID = os.environ.get("UCI_GIST_ID", "").strip()
GIST_TOKEN = os.environ.get("UCI_GIST_TOKEN", "").strip()

# User-Agent réaliste pour passer la détection Instagram
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/121.0.0.0 Safari/537.36"
)

HEADLESS = True  # Passez à False pour voir le navigateur si Instagram bloque

# ═══════════════════════════════════════════════════════════════════
# LISTE DES ENTITÉS (équipes et courses UCI WorldTour 2026)
# Synchronisée avec index.html du dashboard
# ═══════════════════════════════════════════════════════════════════

ENTITIES = [
    # --- Équipes hommes ---
    {"nom": "Alpecin-Premier Tech", "cl": "Belgique", "ig": "alpecin.premiertech"},
    {"nom": "Bahrain Victorious", "cl": "Bahreïn", "ig": "teambahrainvictorious"},
    {"nom": "Decathlon CMA CGM", "cl": "France", "ig": "decathloncmacgmteam"},
    {"nom": "EF Education-Easy Post", "cl": "États-Unis", "ig": "efprocycling"},
    {"nom": "Groupama-FDJ United", "cl": "France", "ig": "groupamafdj"},
    {"nom": "Ineos Grenadiers", "cl": "Royaume-Uni", "ig": "ineosgrenadiers"},
    {"nom": "Lidl-Trek", "cl": "Allemagne", "ig": "lidl_trek"},
    {"nom": "Lotto Intermarché", "cl": "Belgique", "ig": "lotto.cyclingteam"},
    {"nom": "Movistar", "cl": "Espagne", "ig": "movistarteam"},
    {"nom": "NSN", "cl": "Suisse", "ig": "nsncyclingteam"},
    {"nom": "Red Bull-Bora-Hansgrohe", "cl": "Allemagne", "ig": "redbullborahansgrohe"},
    {"nom": "Soudal Quick-Step", "cl": "Belgique", "ig": "soudalquickstepteam"},
    {"nom": "Team Jayco AlUla", "cl": "Australie", "ig": "greenedgecycling"},
    {"nom": "Team Picnic PostNL", "cl": "Pays-Bas", "ig": "teampicnicpostnl"},
    {"nom": "Team Visma-Lease a Bike", "cl": "Pays-Bas", "ig": "teamvisma_leaseabike"},
    {"nom": "UAE Emirates XRG", "cl": "Émirats arabes unis", "ig": "uae_team_emirates"},
    {"nom": "Uno-X Mobility", "cl": "Norvège", "ig": "unoxteam"},
    {"nom": "XDS-Astana", "cl": "Kazakhstan", "ig": "xds_astana_team"},
    # --- Équipes femmes ---
    {"nom": "AG Insurance-Soudal Team", "cl": "Belgique", "ig": "aginsurancesoudal"},
    {"nom": "Canyon//SRAM zondacrypto", "cl": "Allemagne", "ig": "wmncycling"},
    {"nom": "EF Education-Oatly", "cl": "États-Unis", "ig": "efeducationoatly"},
    {"nom": "FDJ United-Suez", "cl": "France", "ig": "fdj.suez"},
    {"nom": "Fenix-Premier Tech", "cl": "Belgique", "ig": "fenix.premiertech"},
    {"nom": "Human Powered Health", "cl": "États-Unis", "ig": "hphcycling"},
    {"nom": "Lidl-Trek", "cl": "Allemagne", "ig": "lidl_trek_women"},
    {"nom": "Liv AlUla Jayco", "cl": "Australie", "ig": "greenedgecycling"},
    {"nom": "Movistar Team", "cl": "Espagne", "ig": "movistarteamwomen"},
    {"nom": "Team Picnic PostNL", "cl": "Pays-Bas", "ig": "teampicnicpostnl"},
    {"nom": "Team SD Worx-Protime", "cl": "Pays-Bas", "ig": "teamsdworxprotime"},
    {"nom": "Team Visma | Lease a Bike", "cl": "Pays-Bas", "ig": "teamvisma_leaseabike_women"},
    {"nom": "UAE Team ADQ", "cl": "Émirats arabes unis", "ig": "uaeteamadq"},
    {"nom": "Uno-X Mobility", "cl": "Norvège", "ig": "unoxteam"},
    # --- Courses hommes ---
    {"nom": "Tour de France", "cl": "2.UWT", "ig": "letourdefrance"},
    {"nom": "Giro d'Italia", "cl": "2.UWT", "ig": "giroditalia"},
    {"nom": "La Vuelta Ciclista a España", "cl": "2.UWT", "ig": "lavuelta"},
    {"nom": "Paris-Roubaix", "cl": "1.UWT", "ig": "parisroubaixcourse"},
    {"nom": "Ronde van Vlaanderen", "cl": "1.UWT", "ig": "rondevanvlaanderen"},
    {"nom": "Milano-Sanremo", "cl": "1.UWT", "ig": "milanosanremo_"},
    {"nom": "La Flèche Wallonne", "cl": "1.UWT", "ig": "classiquesardennes"},
    {"nom": "Liège-Bastogne-Liège", "cl": "1.UWT", "ig": "classiquesardennes"},
    {"nom": "Il Lombardia", "cl": "1.UWT", "ig": "ilombardia"},
    {"nom": "Strade Bianche", "cl": "1.UWT", "ig": "strade_bianche"},
    {"nom": "Paris-Nice", "cl": "2.UWT", "ig": "parisnicecourse"},
    {"nom": "Tirreno-Adriatico", "cl": "2.UWT", "ig": "tirreno_adriatico"},
    {"nom": "Volta a Catalunya", "cl": "2.UWT", "ig": "voltacatalunya"},
    {"nom": "Itzulia Basque Country", "cl": "2.UWT", "ig": "ehitzulia"},
    {"nom": "Tour de Romandie", "cl": "2.UWT", "ig": "tourderomandie"},
    {"nom": "Tour Auvergne-Rhône-Alpes", "cl": "2.UWT", "ig": "criteriumdudauphine"},
    {"nom": "Tour de Suisse", "cl": "2.UWT", "ig": "tourdesuisse_official"},
    {"nom": "Tour de Pologne", "cl": "2.UWT", "ig": "tourdepologne"},
    {"nom": "Renewi Tour", "cl": "2.UWT", "ig": "renewitour"},
    {"nom": "Santos Tour Down Under", "cl": "2.UWT", "ig": "tourdownunder"},
    {"nom": "UAE Tour", "cl": "2.UWT", "ig": "theuaetourofficial"},
    {"nom": "Tour of Guangxi", "cl": "2.UWT", "ig": "tourofguangxi"},
    {"nom": "Amstel Gold Race", "cl": "1.UWT", "ig": "amstelgoldrace"},
    {"nom": "Eschborn-Frankfurt", "cl": "1.UWT", "ig": "eschbornfrankfurt"},
    {"nom": "Tour of Bruges", "cl": "1.UWT", "ig": "bruggedepanne"},
    {"nom": "E3 Saxo Classic", "cl": "1.UWT", "ig": "e3_saxoclassic"},
    {"nom": "In Flanders Fields", "cl": "1.UWT", "ig": "inflandersfieldsrace"},
    {"nom": "Dwars door Vlaanderen", "cl": "1.UWT", "ig": "dwarsdoorvlaanderenofficial"},
    {"nom": "Omloop Nieuwsblad", "cl": "1.UWT", "ig": "omloophetnieuwsbladofficial"},
    {"nom": "Copenhagen Sprint", "cl": "1.UWT", "ig": "cphsprint"},
    {"nom": "Clásica San Sebastián", "cl": "1.UWT", "ig": "dklasikoa"},
    {"nom": "Bretagne Classic", "cl": "1.UWT", "ig": "grandprixplouay"},
    {"nom": "GP Cycliste de Québec", "cl": "1.UWT", "ig": "grandsprixcyclistes"},
    {"nom": "GP Cycliste de Montréal", "cl": "1.UWT", "ig": "grandsprixcyclistes"},
    {"nom": "Cadel Evans Road Race", "cl": "1.UWT", "ig": "cadelroadrace"},
    {"nom": "ADAC Cyclassics", "cl": "1.UWT", "ig": "cyclassics"},
    # --- Courses femmes ---
    {"nom": "Tour de France Femmes", "cl": "2.WWT", "ig": "letourfemmes"},
    {"nom": "Giro d'Italia Women", "cl": "2.WWT", "ig": "giroditaliawomen"},
    {"nom": "Vuelta España Femenina", "cl": "2.WWT", "ig": "lavueltafem"},
    {"nom": "Ronde van Vlaanderen", "cl": "1.WWT", "ig": "rondevanvlaanderen"},
    {"nom": "Paris-Roubaix Femmes", "cl": "1.WWT", "ig": "parisroubaix_femmes"},
    {"nom": "Strade Bianche Donne", "cl": "1.WWT", "ig": "strade_bianche"},
    {"nom": "Amstel Gold Race Ladies", "cl": "1.WWT", "ig": "amstelgoldrace"},
    {"nom": "La Flèche Wallonne Femmes", "cl": "1.WWT", "ig": "classiquesardennes"},
    {"nom": "Liège-Bastogne-Liège Femmes", "cl": "1.WWT", "ig": "classiquesardennes"},
    {"nom": "Santos Tour Down Under", "cl": "2.WWT", "ig": "tourdownunder"},
    {"nom": "UAE Tour Women", "cl": "2.WWT", "ig": "theuaetourofficial"},
    {"nom": "Itzulia Women", "cl": "2.WWT", "ig": "ehitzulia"},
    {"nom": "Vuelta a Burgos Feminas", "cl": "2.WWT", "ig": "vueltaburgos"},
    {"nom": "Tour de Suisse Women", "cl": "2.WWT", "ig": "tourdesuisse_official"},
    {"nom": "Tour of Britain Women", "cl": "2.WWT", "ig": "tourofbritain"},
    {"nom": "Tour de Romandie Féminin", "cl": "2.WWT", "ig": "tourderomandie"},
    {"nom": "Tour of Chongming Island", "cl": "2.WWT", "ig": "tourofchongming"},
    {"nom": "Omloop Nieuwsblad", "cl": "1.WWT", "ig": "omloophetnieuwsbladofficial"},
    {"nom": "Trofeo Alfredo Binda", "cl": "1.WWT", "ig": "trofeobinda"},
    {"nom": "Milano-Sanremo Donne", "cl": "1.WWT", "ig": "milanosanremo_"},
    {"nom": "Tour of Bruges", "cl": "1.WWT", "ig": "bruggedepanne"},
    {"nom": "In Flanders Fields", "cl": "1.WWT", "ig": "inflandersfieldsrace"},
    {"nom": "Dwars door Vlaanderen", "cl": "1.WWT", "ig": "dwarsdoorvlaanderenofficial"},
    {"nom": "Copenhagen Sprint", "cl": "1.WWT", "ig": "cphsprint"},
    {"nom": "Classic Lorient Agglomération", "cl": "1.WWT", "ig": "grandprixplouay"},
    {"nom": "Cadel Evans Road Race", "cl": "1.WWT", "ig": "cadelroadrace"},
    {"nom": "Tour of Guangxi Women", "cl": "1.WWT", "ig": "tourofguangxi"},
]

# ═══════════════════════════════════════════════════════════════════
# SCRAPING
# ═══════════════════════════════════════════════════════════════════

FOLLOWER_REGEX = re.compile(
    r"([\d]+(?:[\.,]\d+)?\s*[KkMm]?)\s*(?:Followers|abonn|seguidores|follower)",
    re.IGNORECASE,
)


def parse_count(text):
    """Convertit '1.5M', '234K', '1,234', '6 500 000' en entier."""
    if text is None:
        return None
    text = text.strip().replace("\u00a0", "").replace(" ", "")
    # Détection des suffixes K/M
    mult = 1
    last = text[-1:].lower() if text else ""
    if last == "k":
        mult = 1_000
        text = text[:-1]
    elif last == "m":
        mult = 1_000_000
        text = text[:-1]
    # Gestion virgule décimale FR vs séparateur milliers EN
    if "." in text and "," in text:
        text = text.replace(",", "")
    elif "," in text and mult > 1:
        text = text.replace(",", ".")
    elif "," in text:
        text = text.replace(",", "")
    try:
        return int(float(text) * mult)
    except ValueError:
        return None


def scrape_instagram(page, username):
    """
    Scrape le nombre de followers d'un compte Instagram.
    Stratégie : lecture de la balise <meta property="og:description">
    qui contient un texte du type "6.5M Followers, 123 Following, 456 Posts".
    """
    url = f"https://www.instagram.com/{username}/"
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        # Laisse le JS hydrater la meta
        page.wait_for_timeout(1200)

        # 1. Meta og:description
        meta = page.query_selector('meta[property="og:description"]')
        if meta:
            content = meta.get_attribute("content") or ""
            m = FOLLOWER_REGEX.search(content)
            if m:
                return parse_count(m.group(1))

        # 2. Meta description (fallback)
        meta = page.query_selector('meta[name="description"]')
        if meta:
            content = meta.get_attribute("content") or ""
            m = FOLLOWER_REGEX.search(content)
            if m:
                return parse_count(m.group(1))

        # 3. Parsing du HTML complet (dernière chance)
        html = page.content()
        m = FOLLOWER_REGEX.search(html)
        if m:
            return parse_count(m.group(1))

        return None

    except Exception as e:
        print(f"    ⚠️  {e}")
        return None


# ═══════════════════════════════════════════════════════════════════
# SAUVEGARDE EXCEL
# ═══════════════════════════════════════════════════════════════════

def save_excel(results):
    """Append les résultats du run dans historique_followers.xlsx."""
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    datetime_str = now.strftime("%Y-%m-%d %H:%M")

    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Instagram"
        ws.append(["Date", "Horodatage", "Nom", "Catégorie/Pays", "Handle IG", "Followers"])

    count = 0
    for r in results:
        if r["followers"] is None:
            continue
        ws.append([date_str, datetime_str, r["nom"], r["cl"], r["ig"], r["followers"]])
        count += 1

    wb.save(EXCEL_FILE)
    print(f"📊 Excel : {count} lignes ajoutées dans {EXCEL_FILE.name}")


# ═══════════════════════════════════════════════════════════════════
# SYNC GIST (optionnel)
# ═══════════════════════════════════════════════════════════════════

def gh_api(url, method="GET", body=None):
    """Appel à l'API GitHub authentifié."""
    headers = {
        "Authorization": f"token {GIST_TOKEN}",
        "Accept": "application/vnd.github+json",
        "User-Agent": "uci-dashboard-tracker",
    }
    data = None
    if body is not None:
        headers["Content-Type"] = "application/json"
        data = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))


def push_to_gist(results):
    """Fusionne les nouvelles données IG dans le Gist public du dashboard."""
    if not GIST_ID or not GIST_TOKEN:
        print("ℹ️  Sync Gist désactivée (UCI_GIST_ID ou UCI_GIST_TOKEN absent)")
        return False

    print(f"☁️  Récupération du Gist {GIST_ID[:8]}...")
    try:
        gist = gh_api(f"https://api.github.com/gists/{GIST_ID}")
    except Exception as e:
        print(f"❌ Lecture Gist échouée : {e}")
        return False

    current = {}
    if GIST_FILENAME in gist.get("files", {}):
        try:
            current = json.loads(gist["files"][GIST_FILENAME]["content"])
        except Exception:
            current = {}

    dt = current.get("dt", {})
    month_key = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug",
                 "sep", "oct", "nov", "dec"][datetime.now().month - 1]

    # Appliquer les nouvelles données IG
    updated_count = 0
    for r in results:
        if r["followers"] is None:
            continue
        key = f"{r['nom']}|{r['cl']}"
        if key not in dt:
            dt[key] = {}
        if month_key not in dt[key]:
            dt[key][month_key] = {}
        dt[key][month_key]["ig"] = r["followers"]
        updated_count += 1

    current["dt"] = dt
    current.setdefault("lastUpdate", {"ig": None, "fb": None, "x": None, "tk": None})
    current["lastUpdate"]["ig"] = datetime.now().strftime("%Y-%m-%d")
    current.setdefault("imports", [])
    current["imports"].append(
        datetime.now().strftime("%Y-%m-%d") +
        f" (Instagram via Python tracker — {updated_count} comptes)"
    )
    current["updatedAt"] = datetime.now(timezone.utc).isoformat()

    body = {
        "files": {
            GIST_FILENAME: {
                "content": json.dumps(current, ensure_ascii=False, indent=2)
            }
        }
    }

    try:
        gh_api(f"https://api.github.com/gists/{GIST_ID}", method="PATCH", body=body)
        print(f"✅ Gist mis à jour : {updated_count} entrées IG pour {month_key}")
        return True
    except Exception as e:
        print(f"❌ Mise à jour Gist échouée : {e}")
        return False


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print(f"🚴 UCI Instagram Tracker — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    # Dédupliquer par handle IG
    unique_handles = {}
    for e in ENTITIES:
        unique_handles.setdefault(e["ig"], []).append(e)

    print(f"📋 {len(ENTITIES)} entités, {len(unique_handles)} comptes IG uniques à scraper\n")

    handle_results = {}  # ig -> followers count
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(user_agent=USER_AGENT, viewport={"width": 1280, "height": 800})
        page = context.new_page()

        for i, handle in enumerate(sorted(unique_handles.keys()), 1):
            print(f"[{i:2d}/{len(unique_handles)}] @{handle:<35} ", end="", flush=True)
            followers = scrape_instagram(page, handle)
            handle_results[handle] = followers
            if followers is not None:
                print(f"✅ {followers:>12,}".replace(",", " "))
            else:
                print("❌ N/A")
            time.sleep(0.8)  # politeness

        browser.close()

    # Construire la liste complète des résultats par entité
    results = []
    for e in ENTITIES:
        results.append({
            "nom": e["nom"],
            "cl": e["cl"],
            "ig": e["ig"],
            "followers": handle_results.get(e["ig"]),
        })

    ok = sum(1 for r in results if r["followers"] is not None)
    fail = len(results) - ok
    print(f"\n🏁 Résultat : {ok} entités OK, {fail} échecs")

    save_excel(results)
    push_to_gist(results)

    print("\n✨ Terminé.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n⏹  Interrompu par l'utilisateur.")
        sys.exit(130)
