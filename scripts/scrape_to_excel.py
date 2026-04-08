#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UCI Social Dashboard — Scraper Excel
=====================================
Script unique qui scrape les followers Instagram, Facebook et X pour toutes
les équipes et courses UCI WorldTour 2026, et écrit les résultats dans 2
fichiers Excel (uci_equipes.xlsx et uci_courses.xlsx).

Chaque fichier a 2 feuilles (Hommes / Femmes) et une colonne par mois.
À chaque run, seule la colonne du mois courant est mise à jour ou créée ;
les colonnes des mois précédents sont préservées — on construit donc un
historique mois après mois.

Pas d'API externe, pas de clé, pas de serveur. 100 % local.

USAGE
=====
    pip install playwright openpyxl
    python -m playwright install chromium
    python scrape_to_excel.py

Les deux fichiers Excel sont créés/mis à jour dans le même dossier que le
script. Vous pouvez ensuite les importer dans le dashboard via le bouton
"Importer Excel".
"""

import re
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    sys.exit("❌ Playwright manquant. Lancez : pip install playwright && python -m playwright install chromium")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("❌ openpyxl manquant. Lancez : pip install openpyxl")


# ═══════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════

SCRIPT_DIR = Path(__file__).resolve().parent
EXCEL_EQUIPES = SCRIPT_DIR / "uci_equipes.xlsx"
EXCEL_COURSES = SCRIPT_DIR / "uci_courses.xlsx"

HEADLESS = True  # Passez à False pour voir le navigateur si besoin de debug
DELAY_BETWEEN_REQUESTS = 0.6  # secondes entre deux profils (politesse)

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/121.0.0.0 Safari/537.36"
)

# Mois en français pour les en-têtes Excel
MOIS_FR = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
           "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]


# ═══════════════════════════════════════════════════════════════════
# ENTITÉS (synchronisées avec index.html)
# Format : (nom, pays_ou_classe, ig_handle, fb_page, x_handle)
# ═══════════════════════════════════════════════════════════════════

TEAMS_MEN = [
    ("Alpecin-Premier Tech", "Belgique", "alpecin.premiertech", "alpecinpremiertech", "AlpecinPT"),
    ("Bahrain Victorious", "Bahreïn", "teambahrainvictorious", "BahrainVictorious", "BHRVictorious"),
    ("Decathlon CMA CGM", "France", "decathloncmacgmteam", "DecathlonCMACGM", "decathloncmacgm"),
    ("EF Education-Easy Post", "États-Unis", "efprocycling", "EFProCycling", "EFprocycling"),
    ("Groupama-FDJ United", "France", "groupamafdj", "equipecyclistegroupamafdj", "GroupamaFDJ"),
    ("Ineos Grenadiers", "Royaume-Uni", "ineosgrenadiers", "INEOSGrenadiers", "INEOSGrenadiers"),
    ("Lidl-Trek", "Allemagne", "lidl_trek", "lidltrekteam", "LidlTrek"),
    ("Lotto Intermarché", "Belgique", "lotto.cyclingteam", "LottoCyclingTeam", "lottocycling_"),
    ("Movistar", "Espagne", "movistarteam", "movistarteam", "Movistar_Team"),
    ("NSN", "Suisse", "nsncyclingteam", "NSNCyclingTeam", "NSNCyclingTeam"),
    ("Red Bull-Bora-Hansgrohe", "Allemagne", "redbullborahansgrohe", "redbullBORAhansgrohe", "RBH_ProCycling"),
    ("Soudal Quick-Step", "Belgique", "soudalquickstepteam", "soudalquickstepteam", "soudalquickstep"),
    ("Team Jayco AlUla", "Australie", "greenedgecycling", "GreenEDGECycling", "GreenEDGEteam"),
    ("Team Picnic PostNL", "Pays-Bas", "teampicnicpostnl", "teampicnicpostnl", "picnicpostnl"),
    ("Team Visma-Lease a Bike", "Pays-Bas", "teamvisma_leaseabike", "vismaleaseabike", "vismaleaseabike"),
    ("UAE Emirates XRG", "Émirats arabes unis", "uae_team_emirates", "UAETEAMEMIRATES", "TeamEmiratesUAE"),
    ("Uno-X Mobility", "Norvège", "unoxteam", "unoxteam", "UnoXteam"),
    ("XDS-Astana", "Kazakhstan", "xds_astana_team", "XDSAstanaTeam", "XDSAstanaTeam"),
]

TEAMS_WOMEN = [
    ("AG Insurance-Soudal Team", "Belgique", "aginsurancesoudal", "aginsurancesoudalteam", "AGSoudalTeam"),
    ("Canyon//SRAM zondacrypto", "Allemagne", "wmncycling", "WMNcycling", "CanyonSRAM"),
    ("EF Education-Oatly", "États-Unis", "efeducationoatly", "efeducation.oatly", "EFprocycling"),
    ("FDJ United-Suez", "France", "fdj.suez", "FDJUNITEDSUEZ", "FDJSuez"),
    ("Fenix-Premier Tech", "Belgique", "fenix.premiertech", "fenixpremiertech", "FenixPT"),
    ("Human Powered Health", "États-Unis", "hphcycling", "HumanPoweredHealthCycling", "HPHCycling"),
    ("Lidl-Trek", "Allemagne", "lidl_trek_women", "lidltrekteam", "LidlTrekWomen"),
    ("Liv AlUla Jayco", "Australie", "greenedgecycling", "GreenEDGECycling", "GreenEDGEteam"),
    ("Movistar Team", "Espagne", "movistarteamwomen", "movistarteam", "Movistar_Team"),
    ("Team Picnic PostNL", "Pays-Bas", "teampicnicpostnl", "teampicnicpostnl", "picnicpostnl"),
    ("Team SD Worx-Protime", "Pays-Bas", "teamsdworxprotime", "teamsdworx", "SDWorxProtime"),
    ("Team Visma | Lease a Bike", "Pays-Bas", "teamvisma_leaseabike_women", "vismaleaseabike", "vismaleaseabike"),
    ("UAE Team ADQ", "Émirats arabes unis", "uaeteamadq", "uaeteamadq", "UAETeamADQ"),
    ("Uno-X Mobility", "Norvège", "unoxteam", "unoxteam", "UnoXteam"),
]

RACES_MEN = [
    ("Tour de France", "2.UWT", "letourdefrance", "letour", "LeTour"),
    ("Giro d'Italia", "2.UWT", "giroditalia", "giroditalia", "giroditalia"),
    ("La Vuelta Ciclista a España", "2.UWT", "lavuelta", "lavuelta", "lavuelta"),
    ("Paris-Roubaix", "1.UWT", "parisroubaixcourse", "ParisRoubaix", "Paris_Roubaix"),
    ("Ronde van Vlaanderen", "1.UWT", "rondevanvlaanderen", "rondevanvlaanderen", "RondeVlaanderen"),
    ("Milano-Sanremo", "1.UWT", "milanosanremo_", "misanremo", "Milano_Sanremo"),
    ("La Flèche Wallonne", "1.UWT", "classiquesardennes", "FlecheWallonne", "ClassArdennes"),
    ("Liège-Bastogne-Liège", "1.UWT", "classiquesardennes", "liegebastogneliege", "ClassArdennes"),
    ("Il Lombardia", "1.UWT", "ilombardia", "il.lombardia", "Il_Lombardia"),
    ("Strade Bianche", "1.UWT", "strade_bianche", "stradebianche", "Strade_Bianche"),
    ("Paris-Nice", "2.UWT", "parisnicecourse", "parisnicecourse", "parisnice"),
    ("Tirreno-Adriatico", "2.UWT", "tirreno_adriatico", "tirrenoadriatico", "TirrenoAdriatco"),
    ("Volta a Catalunya", "2.UWT", "voltacatalunya", "VoltaCatalunya", "VoltaCatalunya"),
    ("Itzulia Basque Country", "2.UWT", "ehitzulia", "ehitzulia", "ehitzulia"),
    ("Tour de Romandie", "2.UWT", "tourderomandie", "tourderomandie", "TourDeRomandie"),
    ("Tour Auvergne-Rhône-Alpes", "2.UWT", "criteriumdudauphine", "CriteriumDuDauphine", "dauphine"),
    ("Tour de Suisse", "2.UWT", "tourdesuisse_official", "tourdesuisse", "tds"),
    ("Tour de Pologne", "2.UWT", "tourdepologne", "tourdepologne", "tourdepologne"),
    ("Renewi Tour", "2.UWT", "renewitour", "renewitour", "RenewiTour"),
    ("Santos Tour Down Under", "2.UWT", "tourdownunder", "tourdownunder", "tourdownunder"),
    ("UAE Tour", "2.UWT", "theuaetourofficial", "uaetourofficial", "uae_tour"),
    ("Tour of Guangxi", "2.UWT", "tourofguangxi", "TourofGuangxi", "TourofGuangxi"),
    ("Amstel Gold Race", "1.UWT", "amstelgoldrace", "amstelgoldrace", "amstelgoldrace"),
    ("Eschborn-Frankfurt", "1.UWT", "eschbornfrankfurt", "eschbornfrankfurt", "EschbornFrankfurt"),
    ("Tour of Bruges", "1.UWT", "bruggedepanne", "bruggedepanne", "BruggeClassic"),
    ("E3 Saxo Classic", "1.UWT", "e3_saxoclassic", "E3SaxoClassic", "E3SaxoClassic"),
    ("In Flanders Fields", "1.UWT", "inflandersfieldsrace", "gentwevelgemofficial", "IFF_race"),
    ("Dwars door Vlaanderen", "1.UWT", "dwarsdoorvlaanderenofficial", "DwarsDoorVlaanderen", "DwarsDVlaanderen"),
    ("Omloop Nieuwsblad", "1.UWT", "omloophetnieuwsbladofficial", "omloophetnieuwsblad", "OmloopHNB"),
    ("Copenhagen Sprint", "1.UWT", "cphsprint", "copenhagensprint", "cphsprint"),
    ("Clásica San Sebastián", "1.UWT", "dklasikoa", "sansebastianklasikoah", "ClassicSanSeb"),
    ("Bretagne Classic", "1.UWT", "grandprixplouay", "GrandPrixPlouay", "GrandPrixPlouay"),
    ("GP Cycliste de Québec", "1.UWT", "grandsprixcyclistes", "GPCQM", "GPCQM"),
    ("GP Cycliste de Montréal", "1.UWT", "grandsprixcyclistes", "GPCQM", "GPCQM"),
    ("Cadel Evans Road Race", "1.UWT", "cadelroadrace", "CadelRoadRace", "cadelroadrace"),
    ("ADAC Cyclassics", "1.UWT", "cyclassics", "Cyclassics", ""),
]

RACES_WOMEN = [
    ("Tour de France Femmes", "2.WWT", "letourfemmes", "letourfemmes", "LeTourFemmes"),
    ("Giro d'Italia Women", "2.WWT", "giroditaliawomen", "giroditaliawomen", "giroditaliawomen"),
    ("Vuelta España Femenina", "2.WWT", "lavueltafem", "lavuelta", "lavuelta"),
    ("Ronde van Vlaanderen", "1.WWT", "rondevanvlaanderen", "rondevanvlaanderen", "RondeVlaanderen"),
    ("Paris-Roubaix Femmes", "1.WWT", "parisroubaix_femmes", "ParisRoubaix", "Paris_Roubaix"),
    ("Strade Bianche Donne", "1.WWT", "strade_bianche", "stradebianche", "Strade_Bianche"),
    ("Amstel Gold Race Ladies", "1.WWT", "amstelgoldrace", "amstelgoldrace", "amstelgoldrace"),
    ("La Flèche Wallonne Femmes", "1.WWT", "classiquesardennes", "FlecheWallonne", "ClassArdennes"),
    ("Liège-Bastogne-Liège Femmes", "1.WWT", "classiquesardennes", "liegebastogneliege", "ClassArdennes"),
    ("Santos Tour Down Under", "2.WWT", "tourdownunder", "tourdownunder", "tourdownunder"),
    ("UAE Tour Women", "2.WWT", "theuaetourofficial", "uaetourofficial", "uae_tour"),
    ("Itzulia Women", "2.WWT", "ehitzulia", "ehitzulia", "ehitzulia"),
    ("Vuelta a Burgos Feminas", "2.WWT", "vueltaburgos", "VueltaBurgos", "VueltaBurgos"),
    ("Tour de Suisse Women", "2.WWT", "tourdesuisse_official", "tourdesuisse", "tds"),
    ("Tour of Britain Women", "2.WWT", "tourofbritain", "TourofBritain", "TourofBritain"),
    ("Tour de Romandie Féminin", "2.WWT", "tourderomandie", "tourderomandie", "TourDeRomandie"),
    ("Tour of Chongming Island", "2.WWT", "tourofchongming", "TourofChongming", "TourofChongming"),
    ("Omloop Nieuwsblad", "1.WWT", "omloophetnieuwsbladofficial", "omloophetnieuwsblad", "OmloopHNB"),
    ("Trofeo Alfredo Binda", "1.WWT", "trofeobinda", "trofeobinda", "TrofeoBinda"),
    ("Milano-Sanremo Donne", "1.WWT", "milanosanremo_", "misanremo", "Milano_Sanremo"),
    ("Tour of Bruges", "1.WWT", "bruggedepanne", "bruggedepanne", "BruggeClassic"),
    ("In Flanders Fields", "1.WWT", "inflandersfieldsrace", "gentwevelgemofficial", "IFF_race"),
    ("Dwars door Vlaanderen", "1.WWT", "dwarsdoorvlaanderenofficial", "DwarsDoorVlaanderen", "DwarsDVlaanderen"),
    ("Copenhagen Sprint", "1.WWT", "cphsprint", "copenhagensprint", "cphsprint"),
    ("Classic Lorient Agglomération", "1.WWT", "grandprixplouay", "GrandPrixPlouay", "GrandPrixPlouay"),
    ("Cadel Evans Road Race", "1.WWT", "cadelroadrace", "CadelRoadRace", "cadelroadrace"),
    ("Tour of Guangxi Women", "1.WWT", "tourofguangxi", "TourofGuangxi", "TourofGuangxi"),
]


# ═══════════════════════════════════════════════════════════════════
# SCRAPING — INSTAGRAM
# ═══════════════════════════════════════════════════════════════════

def scrape_instagram(page, handle):
    """Renvoie le nombre EXACT de followers Instagram, ou None."""
    if not handle:
        return None
    url = f"https://www.instagram.com/{handle}/"
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(1200)
        html = page.content()

        # 1. JSON inline edge_followed_by.count (exact)
        m = re.search(r'"edge_followed_by"\s*:\s*\{\s*"count"\s*:\s*(\d+)', html)
        if m:
            return int(m.group(1))
        # 2. JSON inline follower_count (exact)
        m = re.search(r'"follower_count"\s*:\s*(\d+)', html)
        if m:
            return int(m.group(1))
        # 3. JSON-LD userInteractionCount (exact)
        m = re.search(r'"userInteractionCount"\s*:\s*"?(\d+)"?', html)
        if m:
            return int(m.group(1))
        return None
    except Exception as e:
        print(f"    ⚠️  IG {handle}: {e}")
        return None


# ═══════════════════════════════════════════════════════════════════
# SCRAPING — FACEBOOK
# ═══════════════════════════════════════════════════════════════════

def scrape_facebook(page, page_slug):
    """Renvoie le nombre EXACT de followers Facebook, ou None."""
    if not page_slug:
        return None
    url = f"https://www.facebook.com/{page_slug}/"
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(1800)
        html = page.content()

        # Stratégie 1 — "follower_count":N dans le JSON embarqué (exact)
        m = re.search(r'"follower_count"\s*:\s*(\d+)', html)
        if m:
            return int(m.group(1))
        # Stratégie 2 — "followerCount":N
        m = re.search(r'"followerCount"\s*:\s*(\d+)', html)
        if m:
            return int(m.group(1))
        # Stratégie 3 — meta description : "123,456 followers"
        m = re.search(r'<meta[^>]+name="description"[^>]+content="([^"]+)"', html)
        if m:
            desc = m.group(1)
            m2 = re.search(r'([\d,\s\u00a0\.]+)\s*(?:people follow|followers|abonnés|personnes suivent|personas siguen|personas seguidoras)', desc, re.I)
            if m2:
                digits = re.sub(r'[^\d]', '', m2.group(1))
                if digits:
                    return int(digits)
        # Stratégie 4 — "X people follow this" dans le HTML visible
        m = re.search(r'([\d,\s\u00a0\.]+)\s*(?:people follow this|followers|abonnés)', html)
        if m:
            digits = re.sub(r'[^\d]', '', m.group(1))
            if digits and len(digits) >= 3:
                return int(digits)
        return None
    except Exception as e:
        print(f"    ⚠️  FB {page_slug}: {e}")
        return None


# ═══════════════════════════════════════════════════════════════════
# SCRAPING — X (Twitter)
# ═══════════════════════════════════════════════════════════════════

def scrape_x(page, handle):
    """Renvoie le nombre EXACT de followers X (via nitter.net d'abord, puis x.com en fallback)."""
    if not handle:
        return None
    # 1. Nitter mirror (pas de login, plus fiable pour scraping)
    nitter_mirrors = ["nitter.net", "nitter.privacydev.net", "nitter.poast.org"]
    for mirror in nitter_mirrors:
        try:
            url = f"https://{mirror}/{handle}"
            page.goto(url, wait_until="domcontentloaded", timeout=20000)
            page.wait_for_timeout(800)
            html = page.content()
            m = re.search(r'<span class="profile-stat-header">Followers</span>\s*<span class="profile-stat-num">([\d,\.]+)', html)
            if m:
                return int(re.sub(r'[^\d]', '', m.group(1)))
            m = re.search(r'title="([\d,]+)"\s*>\s*[\d,\.]+[KM]?\s*</[^>]+>\s*<[^>]+>\s*Followers', html)
            if m:
                return int(re.sub(r'[^\d]', '', m.group(1)))
        except Exception:
            continue

    # 2. Fallback : x.com directement (souvent bloqué mais on tente)
    try:
        url = f"https://x.com/{handle}"
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(2500)
        html = page.content()
        m = re.search(r'"followers_count"\s*:\s*(\d+)', html)
        if m:
            return int(m.group(1))
        # Data from JSON-LD
        m = re.search(r'"interactionStatistic"[^}]*"userInteractionCount"\s*:\s*"?(\d+)"?', html)
        if m:
            return int(m.group(1))
    except Exception as e:
        print(f"    ⚠️  X {handle}: {e}")
    return None


# ═══════════════════════════════════════════════════════════════════
# EXCEL — LECTURE / ÉCRITURE
# ═══════════════════════════════════════════════════════════════════

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="27272A")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def ensure_sheet_structure(wb, sheet_name, entities, month_label):
    """
    S'assure que la feuille a l'en-tête + une ligne par (entité × réseau),
    et renvoie le numéro de la colonne à remplir pour ce mois.
    """
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Si la feuille est vide, on écrit les en-têtes et les lignes des entités
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        # En-tête
        headers = ["Nom", "Pays/Classe", "Réseau"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = HEADER_ALIGN
        row = 2
        for nom, cl, ig, fb, x in entities:
            for net in ("Instagram", "Facebook", "X (Twitter)"):
                ws.cell(row=row, column=1, value=nom)
                ws.cell(row=row, column=2, value=cl)
                ws.cell(row=row, column=3, value=net)
                row += 1
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 14
        ws.freeze_panes = "D2"

    # Garantir que chaque entité × réseau existe (ajouter celles qui manquent)
    existing = set()
    for r in range(2, ws.max_row + 1):
        nom = ws.cell(r, 1).value
        cl = ws.cell(r, 2).value
        net = ws.cell(r, 3).value
        if nom:
            existing.add((nom, cl, net))
    next_row = ws.max_row + 1
    for nom, cl, ig, fb, x in entities:
        for net in ("Instagram", "Facebook", "X (Twitter)"):
            if (nom, cl, net) not in existing:
                ws.cell(row=next_row, column=1, value=nom)
                ws.cell(row=next_row, column=2, value=cl)
                ws.cell(row=next_row, column=3, value=net)
                next_row += 1

    # Trouver ou créer la colonne du mois courant
    month_col = None
    for col in range(4, ws.max_column + 2):
        v = ws.cell(1, col).value
        if v == month_label:
            month_col = col
            break
        if v is None:
            ws.cell(row=1, column=col, value=month_label)
            c = ws.cell(row=1, column=col)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = HEADER_ALIGN
            ws.column_dimensions[get_column_letter(col)].width = 14
            month_col = col
            break

    return ws, month_col


def write_values(ws, month_col, entities, results):
    """
    Écrit les valeurs scrapées dans la colonne du mois courant.
    `results` est un dict : {(nom, cl, net) -> int | None}
    """
    # Indexer les lignes par (nom, cl, net)
    row_index = {}
    for r in range(2, ws.max_row + 1):
        nom = ws.cell(r, 1).value
        cl = ws.cell(r, 2).value
        net = ws.cell(r, 3).value
        if nom:
            row_index[(nom, cl, net)] = r

    for (nom, cl, net), val in results.items():
        r = row_index.get((nom, cl, net))
        if not r:
            continue
        ws.cell(row=r, column=month_col, value=val if val is not None else "")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

def process_entity_set(page, entities, label):
    """Scrape tous les réseaux pour une liste d'entités. Renvoie un dict de résultats."""
    results = {}  # (nom, cl, net) -> int | None
    total = len(entities)
    print(f"\n🔹 {label} ({total} entités)")
    for i, (nom, cl, ig, fb, x) in enumerate(entities, 1):
        print(f"  [{i:2d}/{total}] {nom[:38]:<38}", end="", flush=True)

        # Instagram
        ig_val = scrape_instagram(page, ig) if ig else None
        time.sleep(DELAY_BETWEEN_REQUESTS)

        # Facebook
        fb_val = scrape_facebook(page, fb) if fb else None
        time.sleep(DELAY_BETWEEN_REQUESTS)

        # X
        x_val = scrape_x(page, x) if x else None
        time.sleep(DELAY_BETWEEN_REQUESTS)

        results[(nom, cl, "Instagram")] = ig_val
        results[(nom, cl, "Facebook")] = fb_val
        results[(nom, cl, "X (Twitter)")] = x_val

        def fmt(v):
            return f"{v:>10,}".replace(",", " ") if v is not None else "       N/A"
        print(f" IG {fmt(ig_val)}  FB {fmt(fb_val)}  X {fmt(x_val)}")

    return results


def update_excel(path, hommes_entities, femmes_entities, hommes_results, femmes_results, month_label):
    """Ouvre ou crée un classeur et y écrit les résultats dans les 2 feuilles."""
    if path.exists():
        wb = load_workbook(path)
        # Supprimer la feuille par défaut "Sheet" si présente
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]
    else:
        wb = Workbook()
        # Supprimer la feuille créée par défaut
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws_h, col_h = ensure_sheet_structure(wb, "Hommes", hommes_entities, month_label)
    write_values(ws_h, col_h, hommes_entities, hommes_results)

    ws_f, col_f = ensure_sheet_structure(wb, "Femmes", femmes_entities, month_label)
    write_values(ws_f, col_f, femmes_entities, femmes_results)

    wb.save(path)
    print(f"💾 {path.name} mis à jour (colonne « {month_label} »)")


def main():
    now = datetime.now()
    month_label = f"{MOIS_FR[now.month - 1]} {now.year}"

    print("=" * 72)
    print(f"🚴  UCI Social Dashboard — Scraper Excel")
    print(f"📅 Mois courant : {month_label}")
    print("=" * 72)

    total_entities = len(TEAMS_MEN) + len(TEAMS_WOMEN) + len(RACES_MEN) + len(RACES_WOMEN)
    print(f"📋 {total_entities} entités à scraper × 3 réseaux = {total_entities * 3} requêtes\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(
            user_agent=USER_AGENT,
            viewport={"width": 1280, "height": 800},
            locale="fr-FR",
        )
        page = context.new_page()

        teams_men_res = process_entity_set(page, TEAMS_MEN, "Équipes hommes")
        teams_women_res = process_entity_set(page, TEAMS_WOMEN, "Équipes femmes")
        races_men_res = process_entity_set(page, RACES_MEN, "Courses hommes")
        races_women_res = process_entity_set(page, RACES_WOMEN, "Courses femmes")

        browser.close()

    print("\n📊 Écriture des fichiers Excel...")
    update_excel(EXCEL_EQUIPES, TEAMS_MEN, TEAMS_WOMEN, teams_men_res, teams_women_res, month_label)
    update_excel(EXCEL_COURSES, RACES_MEN, RACES_WOMEN, races_men_res, races_women_res, month_label)

    # Statistiques
    all_results = {}
    for r in (teams_men_res, teams_women_res, races_men_res, races_women_res):
        all_results.update(r)
    ok = sum(1 for v in all_results.values() if v is not None)
    total = len(all_results)
    pct = (ok / total * 100) if total else 0
    print(f"\n✨ Terminé : {ok}/{total} chiffres récupérés ({pct:.0f}%)")
    print(f"   Importez {EXCEL_EQUIPES.name} et {EXCEL_COURSES.name} dans le dashboard.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n⏹  Interrompu.")
        sys.exit(130)
